from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
wb.create_sheet("Sheet_one")
wb.create_sheet("Sheet_two")

wb['Sheet'].title = 'Sheet_zero'




ws0 = wb['Sheet_zero']

wb.copy_worksheet(ws0)

ws1 = wb['Sheet_one']

ws1['B2'] = 'Fruits'
ws1['C2'] = 'Sales'

wb.save('book_eg.xlsx')


wb2 = load_workbook(filename = 'testbook.xlsx')

#sheet_ranges = wb['Sheet1']
#print(sheet_ranges['A1'].value)

wb2.save('book_eg.xlsx')